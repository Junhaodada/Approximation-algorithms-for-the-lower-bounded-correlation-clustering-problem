import pandas as pd
import random
import copy
import csv
from typing import List
import openpyxl


def zuhe(arr):
    """返回划分数组的二元组,构造r矩阵"""
    res = []
    for i in arr:
        for j in i:
            for k in i:
                if j != k:
                    res.append([j, k])
    return res


def zuhe2(arr):
    res = []
    for i in arr:
        for j in arr:
            res.append([i, j])
    return res


class Graph:
    def __init__(self, num, lower) -> None:
        self.vetex = []
        self.edge = []
        self.n = num
        self.L = lower

    def initG1(self):
        for i in range(self.n):
            self.edge.append([-1 for i in range(self.n)])
        for i in range(self.n):
            self.edge[i][i] = 0
        self.vetex = [i for i in range(self.n)]

    def createG1(self):
        self.initG1()
        self.edge[0][1] = 1
        self.edge[1][0] = 1
        # self.edge[2][3] = 1
        # self.edge[3][2] = 1
        # self.edge[4][5] = 1
        # self.edge[5][4] = 1
        # self.edge[6][7] = 1
        # self.edge[7][6] = 1
        # self.edge[8][9] = 1
        # self.edge[9][8] = 1
        self.printG()

    def printG(self):
        edge_file = openpyxl.load_workbook('edge.xlsx')
        sheet = edge_file['Sheet']
        for i in range(100):
            for j in range(100):
                sheet.cell(i+1, j+1).value = None
        for i in range(self.n):
            for j in range(self.n):
                sheet.cell(i+1, j+1).value = self.edge[i][j]
        # with open(filename, 'w', encoding='utf8', newline='') as f:
        #     writer = csv.writer(f)
        #     writer.writerows(self.edge)
        # print('写入文件成功!')
        edge_file.save('edge.xlsx')
        edge_file.close()
        # for i in range(len(self.edge)):
        #     for j in range(len(self.edge[i])):
        #         if(j == len(self.edge[i])-1):
        #             print(self.edge[i][j])
        #         else:
        #             print(self.edge[i][j], ',', end='')

    def algo1(self):
        U = self.vetex
        Part = []
        while len(U) != 0:
            arr = []
            if len(U) >= 2*self.L:
                for i in range(self.L):
                    r = random.randint(0, len(U)-1)
                    arr.append(U[r])
                    U.remove(U[r])
            else:
                # arr = U
                arr = copy.deepcopy(U)
                U.clear()
            Part.append(arr)
        r = []
        for i in range(self.n):
            r.append([1 for j in range(self.n)])
        for i in range(self.n):
            r[i][i] = 0
        for i in zuhe(Part):
            r[i[0]][i[1]] = 0
        return Part, r

    def initG2(self):
        for i in range(self.n):
            self.edge.append([1 for i in range(self.n)])
        for i in range(self.n):
            self.edge[i][i] = 0
        self.vetex = [i for i in range(self.n)]

    def createG2(self):
        self.initG2()
        self.edge[0][1] = -1
        self.edge[1][0] = -1
        self.edge[0][3] = -1
        self.edge[3][0] = -1
        self.edge[2][3] = -1
        self.edge[3][2] = -1
        self.edge[4][5] = -1
        self.edge[5][4] = -1
        self.edge[6][7] = -1
        self.edge[7][6] = -1
        self.edge[8][9] = -1
        self.edge[9][8] = -1

        # self.edge[4][3] = -1
        # self.edge[3][4] = -1
        # self.edge[10][5] = -1
        # self.edge[5][10] = -1
        # self.edge[11][7] = -1
        # self.edge[7][11] = -1
        # self.edge[8][18] = -1
        # self.edge[18][8] = -1
        # self.edge[6][13] = -1
        # self.edge[6][13] = -1

    def algo2(self, r_lp: List):
        U = self.vetex
        C = set()
        Part = []
        while len(U) != 0:
            Tq1 = []
            Tq2 = []
            Tq_all = {}
            for q in U:
                p_r = {}
                for p in U:
                    p_r.setdefault(p, r_lp[q][p])
                tmp = sorted(p_r.items(), key=lambda x: x[1])
                for i in range(self.L):
                    Tq1.append(tmp[:self.L][i][0])
                for t in U:
                    if r_lp[q][t] <= 0.5:
                        Tq2.append(t)
                Tq = set(Tq1).union(set(Tq2))
                # print(Tq1, Tq2, Tq)
                Tq_all.setdefault(q, (list(Tq)))
            ave_Tq = {}
            sum = 0
            for k, tq in Tq_all.items():
                for i in tq:
                    sum += r_lp[k][i]
                ave_Tq.setdefault(k, sum/len(tq))
            q_min = min(ave_Tq.items(), key=lambda x: x[1])
            U_cpy = copy.deepcopy(U)
            for i in Tq_all[q_min[0]]:
                U_cpy.remove(i)
            if q_min[1] <= 1/20 and len(U_cpy) >= self.L:
                Part.append(Tq_all[q_min[0]])
                for i in Tq_all[q_min[0]]:
                    U.remove(i)
                C.add(q_min[0])
            else:
                U_cpy2 = copy.deepcopy(U)
                Part.append(U_cpy2)
                U.clear()
                C.add(q_min[0])
        r = []
        for i in range(self.n):
            r.append([1 for j in range(self.n)])
        for i in range(self.n):
            r[i][i] = 0
        for i in zuhe(Part):
            r[i[0]][i[1]] = 0
        return Part, r, C

    def CreateRandomG(self, b):
        # 每个点的负边个数小于(n-1)/5
        # n 20
        # supper 3
        self.vetex = [i for i in range(self.n)]
        for i in range(self.n):
            self.edge.append([0 for k in range(self.n)])
            for j in range(self.n):
                # random.seed(j)
                r_t = random.randint(0, 1)
                if r_t == 1:
                    self.edge[i][j] = 1
                else:
                    self.edge[i][j] = -1
        for i in range(self.n):
            self.edge[i][i] = 0
        for i in range(self.n):
            t = 0
            while self.edge[i].count(-1) > b:
                t = self.edge[i].index(-1, t+1, -1)
                self.edge[i][t] = 1
                self.edge[t][i] = 1
            for j in range(self.n):
                self.edge[j][i] = self.edge[i][j]
        self.printG()

    def CreateRandomG2(self):
        self.vetex = [i for i in range(self.n)]
        for i in range(self.n):
            self.edge.append([0 for k in range(self.n)])
        for i in range(self.n-1):
            for j in range(i, self.n):
                r_t = random.randint(0, 1)
                if r_t == 1:
                    self.edge[i][j] = 1
                    self.edge[j][i] = 1
                else:
                    self.edge[i][j] = -1
                    self.edge[j][i] = -1
        for i in range(self.n):
            self.edge[i][i] = 0
        self.printG()

    def algo3(self, r_lp: List):
        U = self.vetex
        C = set()
        Part = []
        Tq1 = []
        Tq2 = []
        Tq_all = {}
        for q in U:
            p_r = {}
            for p in U:
                p_r.setdefault(p, r_lp[q][p])
            tmp = sorted(p_r.items(), key=lambda x: x[1])
            l = max(self.L, 2*(self.n-1)//5)
            for i in range(l):
                Tq1.append(tmp[:l][i][0])
            for t in U:
                if r_lp[q][t] <= 0.5:
                    Tq2.append(t)
            Tq = set(Tq1).union(set(Tq2))
            # print(Tq1, Tq2, Tq)
            Tq_all.setdefault(q, (list(Tq)))
        ave_Tq = {}
        sum = 0
        for k, tq in Tq_all.items():
            for i in tq:
                sum += r_lp[k][i]
            ave_Tq.setdefault(k, sum/len(tq))
        q_min = min(ave_Tq.items(), key=lambda x: x[1])
        U_cpy = copy.deepcopy(U)
        for i in Tq_all[q_min[0]]:
            U_cpy.remove(i)
        if q_min[1] <= 17/80 and len(U_cpy) >= self.L:
            Part.append(Tq_all[q_min[0]])
            for i in Tq_all[q_min[0]]:
                U.remove(i)
            Part.append(U)
            C.add(q_min[0])
        else:
            U_cpy2 = copy.deepcopy(U)
            Part.append(U_cpy2)
            U.clear()
            C.add(q_min[0])
        r = []
        for i in range(self.n):
            r.append([1 for j in range(self.n)])
        for i in range(self.n):
            r[i][i] = 0
        for i in zuhe(Part):
            r[i[0]][i[1]] = 0
        return Part, r, C


def test2(n, L, tag=2):
    # 算法2测试
    # 初始化图
    G = Graph(n, L)
    G.initG2()
    # 从文件中读取edge
    g_file = openpyxl.load_workbook('edge.xlsx')
    sheet = g_file['Sheet']
    for i in range(n):
        for j in range(n):
            G.edge[i][j] = sheet.cell(i+1, j+1).value
    g_file.close()
    # 初始化r_lp
    r_lp = []
    for i in range(n):
        r_lp.append([1 for j in range(n)])
    for i in range(n):
        r_lp[i][i] = 0
    r_lp_file = openpyxl.load_workbook('r.xlsx')
    sheet = r_lp_file['Sheet']
    for i in range(n):
        for j in range(n):
            r_lp[i][j] = sheet.cell(i+1, j+1).value
    r_lp_file.close()
    # 算法3
    if tag == 2:
        part, r, C = G.algo2(r_lp)
    elif tag == 3:
        part, r, C = G.algo3(r_lp)
    print('算法2划分和聚点:\n', part, C)
    print('r矩阵:')
    for i in range(len(G.edge)):
        for j in range(len(G.edge[i])):
            if(j == len(G.edge[i])-1):
                print(r[i][j])
            else:
                print(r[i][j], ',', end='')
    # 算法2的目标函数计算
    obj = 0
    for i in range(G.n):
        for j in range(G.n):
            if G.edge[i][j] > 0:
                obj += r[i][j]
            elif G.edge[i][j] < 0:
                obj += (1-r[i][j])
    print('obj2:\n', obj)


def test1(n, L):
    # 算法1测试
    G = Graph(n, L)
    G.createG1()
    # print('邻接矩阵:')
    G.printG()
    part, r = G.algo1()
    print('算法1划分:\n', part)
    print('r矩阵:')
    for i in range(len(G.edge)):
        for j in range(len(G.edge[i])):
            if(j == len(G.edge[i])-1):
                print(r[i][j])
            else:
                print(r[i][j], ',', end='')
    print()

    # 算法1的目标函数计算
    obj = 0
    for i in range(G.n):
        for j in range(G.n):
            if G.edge[i][j] > 0:
                obj += r[i][j]
            elif G.edge[i][j] < 0:
                obj += (1-r[i][j])
    print('obj1:', obj)

